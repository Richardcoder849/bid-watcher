#!/usr/bin/env python3
"""
周报生成脚本
支持 Markdown 和 Excel 格式的投标情报周报
"""

import json
import os
import shutil
from datetime import datetime
import glob
import sys

# 添加 UTF-8 支持
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# 尝试导入 openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("警告: openpyxl 未安装，将只生成 Markdown 格式")

# 基于脚本自身位置确定 data 目录
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(os.path.dirname(SCRIPT_DIR), "data")

COMPETITORS = {
    "先导智能": "无锡先导智能装备股份有限公司",
    "海目星": "海目星激光科技集团股份有限公司",
    "赢合科技": "深圳市赢合科技股份有限公司",
    "联赢激光": "深圳市联赢激光股份有限公司"
}

def load_latest_enriched():
    """加载最新补充背景后的数据"""
    import glob
    files = glob.glob(os.path.join(DATA_DIR, "bids_parsed_*enriched*.json"))
    if not files:
        files = glob.glob(os.path.join(DATA_DIR, "bids_raw_*.json"))
    if not files:
        return []
    latest = max(files)
    print(f"加载数据: {latest}")
    with open(latest, 'r', encoding='utf-8') as f:
        data = json.load(f)
        if isinstance(data, dict) and 'bids' in data:
            return data.get('bids', [])
        return data if isinstance(data, list) else []

def archive_to_history(bids, week_num, year):
    """存档本周数据到 history/"""
    history_dir = os.path.join(DATA_DIR, "history")
    os.makedirs(history_dir, exist_ok=True)

    # 存档 JSON 数据
    json_file = os.path.join(history_dir, f"week_{year}WW{week_num:02d}.json")
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(bids, f, ensure_ascii=False, indent=2)

    # 复制周报 Markdown（如果存在）
    md_files = glob.glob(os.path.join(DATA_DIR, "bid_report_*.md"))
    if md_files:
        latest_md = max(md_files)
        md_archive = os.path.join(history_dir, f"week_{year}WW{week_num:02d}_report.md")
        shutil.copy2(latest_md, md_archive)

    return json_file

def generate_markdown_report(bids, week_num=None):
    """生成 Markdown 格式周报"""

    if week_num is None:
        today = datetime.now()
        week_num = today.isocalendar()[1]
        year = today.year
    else:
        year = datetime.now().year

    today_str = datetime.now().strftime("%Y年%m月%d日 %H:%M")

    # 统计
    total = len(bids)
    by_competitor = {}
    priority_counts = {"S": 0, "A": 0, "B": 0, "C": 0}
    relevance_counts = {"高": 0, "低": 0}
    enriched_count = 0

    for b in bids:
        c = b.get('competitor', '')
        if c:
            by_competitor[c] = by_competitor.get(c, 0) + 1

        p = b.get('优先级', 'C')
        if p in priority_counts:
            priority_counts[p] += 1

        rel = b.get('相关性', '低')
        relevance_counts[rel] = relevance_counts.get(rel, 0) + 1

        if b.get('公司背景') and b.get('公司背景') != '未知（建议人工核实）':
            enriched_count += 1

    # 高优先级招标（S + A）
    high_priority = [b for b in bids if b.get('优先级') in ['S', 'A']]

    report = f"""# 投标情报周报

**第{year}年第{week_num}周**
生成时间：{today_str}
统计来源：{total}条招标线索

---

## 📊 本周概览

| 指标 | 数值 |
|------|------|
| 新增线索总数 | {total} 条 |
| 高相关招标 | {relevance_counts.get('高', 0)} 条 |
| S级 优先跟进 | {priority_counts.get('S', 0)} 条 |
| A级 重点关注 | {priority_counts.get('A', 0)} 条 |
| 已匹配公司背景 | {enriched_count} 条 |

"""

    # 优先级分布
    report += "### 优先级分布\n\n"
    report += "| 优先级 | 数量 | 说明 |\n"
    report += "|--------|------|------|\n"
    report += f"| S | {priority_counts.get('S', 0)} | 高预算 + 高相关，需优先跟进 |\n"
    report += f"| A | {priority_counts.get('A', 0)} | 有明显采购信号，重点关注 |\n"
    report += f"| B | {priority_counts.get('B', 0)} | 行业相关，可跟进 |\n"
    report += f"| C | {priority_counts.get('C', 0)} | 需人工判断 |\n"
    report += "\n"

    # 竞争公司分布
    if by_competitor:
        report += "### 竞争公司投标分布\n\n"
        report += "| 公司 | 相关线索数 |\n"
        report += "|------|-----------|\n"
        for name, cnt in sorted(by_competitor.items(), key=lambda x: -x[1]):
            full_name = COMPETITORS.get(name, name)
            report += f"| {full_name} | {cnt} 条 |\n"
        report += "\n"

    # S/A 级高优先级线索
    if high_priority:
        report += "## 🔥 优先跟进招标（S + A 级）\n\n"
        report += "| # | 公司 | 项目名称 | 预算 | 投标时间 | 优先级 | 链接 |\n"
        report += "|---|------|---------|------|---------|--------|------|\n"

        for i, b in enumerate(high_priority, 1):
            company = b.get('company', '')[:15]
            title = (b.get('title', '') or '未知')[:30]
            amount = b.get('amount', '未知')[:12]
            bid_time = b.get('bid_time', '未知')[:12]
            priority = b.get('优先级', 'C')
            url = b.get('url', '')
            url_short = url[:35] + '...' if len(url) > 35 else url

            report += f"| {i} | {company} | {title} | {amount} | {bid_time} | **{priority}** | [链接]({url}) |\n"
        report += "\n"

    # 全部线索列表
    report += "## 📋 本周招标线索详情\n\n"

    if bids:
        report += "| # | 关键词 | 公司 | 投标时间 | 预算 | 优先级 | 相关性 | 链接 |\n"
        report += "|---|--------|------|---------|------|--------|--------|------|\n"

        for i, b in enumerate(bids, 1):
            keyword = b.get('keyword', '')[:12]
            company = b.get('company', '')[:12]
            bid_time = b.get('bid_time', '未知')[:12]
            amount = b.get('amount', '未知')[:10]
            priority = b.get('优先级', 'C')
            relevance = b.get('相关性', '低')
            url = b.get('url', '')
            url_short = url[:30] + '...' if len(url) > 30 else url

            # 优先级着色：S红色、A橙色、B黄色、C灰
            p_display = priority

            report += f"| {i} | {keyword} | {company} | {bid_time} | {amount} | {p_display} | {relevance} | [链接]({url}) |\n"
    else:
        report += "*本周暂无新增招标线索*\n"

    # 附录
    report += f"""
---

## 📝 附录：监控配置

**监控关键词**: {', '.join(['锂电池', '储能', '装配段', '锂电设备', '电池生产设备'])}

**竞争公司列表**:
"""
    for short, full in COMPETITORS.items():
        report += f"- {full} ({short})\n"

    report += """
**数据来源**: Bing 搜索

**评分规则**: S(≥4分) / A(3分) / B(2分) / C(≤1分)
- 有预算金额 +1，大预算(≥1000万) +1
- 公司在已知库中 +1，有投标时间 +1
- 标题含"储能"/"锂电" +1，竞争公司自身招标 -2

---
*本报告由 bid-watcher Skill 自动生成*
"""

    return report, year

def save_report(markdown_content, year, week_num, output_format='both'):
    """保存周报 - 支持 Markdown 和 Excel 格式"""
    os.makedirs(DATA_DIR, exist_ok=True)
    
    md_filename = os.path.join(DATA_DIR, f"bid_report_W{week_num:02d}_{year}_{datetime.now().strftime('%Y%m%d')}.md")
    
    with open(md_filename, 'w', encoding='utf-8') as f:
        f.write(markdown_content)
    
    print(f"[完成] Markdown周报: {md_filename}")
    
    results = [md_filename]
    
    # 生成 Excel 格式
    if output_format in ['both', 'excel'] and HAS_OPENPYXL:
        excel_file = generate_excel_report(year, week_num)
        if excel_file:
            results.append(excel_file)
    
    return results


def generate_excel_report(year, week_num):
    """生成 Excel 格式周报"""
    if not HAS_OPENPYXL:
        return None
    
    bids = load_latest_enriched()
    if not bids:
        print("警告: 无数据可导出")
        return None
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    
    # 样式
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    S_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    A_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    B_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    C_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Sheet 1: 概览
    ws = wb.active
    ws.title = "本周概览"
    
    # 统计
    total = len(bids)
    priority_counts = {"S": 0, "A": 0, "B": 0, "C": 0}
    for b in bids:
        p = b.get('优先级', 'C')
        priority_counts[p] = priority_counts.get(p, 0) + 1
    
    headers = ["指标", "数值"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    stats_data = [
        ["新增线索总数", total],
        ["S级优先跟进", priority_counts.get('S', 0)],
        ["A级重点关注", priority_counts.get('A', 0)],
        ["B级行业相关", priority_counts.get('B', 0)],
        ["C级需人工判断", priority_counts.get('C', 0)],
    ]
    for row, (label, value) in enumerate(stats_data, 2):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
    
    # Sheet 2: 全部线索
    ws2 = wb.create_sheet("招标线索")
    
    bid_headers = ["序号", "项目名称", "公司", "预算", "投标时间", "优先级", "相关性", "链接"]
    for col, h in enumerate(bid_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    for row, b in enumerate(bids, 2):
        # 兼容不同字段名
        title = b.get('title') or b.get('desc') or b.get('项目名称', '')
        company = b.get('company', '')
        budget = b.get('budget', '')
        deadline = b.get('deadline', b.get('投标时间', ''))
        priority = b.get('优先级', 'C')
        relevance = b.get('相关性', b.get('relevance', ''))
        url = b.get('url', '')
        
        ws2.cell(row=row, column=1, value=row-1)
        ws2.cell(row=row, column=2, value=title[:50])
        ws2.cell(row=row, column=3, value=company)
        ws2.cell(row=row, column=4, value=budget)
        ws2.cell(row=row, column=5, value=deadline)
        
        priority = b.get('优先级', 'C')
        p_cell = ws2.cell(row=row, column=6, value=priority)
        if priority == 'S':
            p_cell.fill = S_FILL
        elif priority == 'A':
            p_cell.fill = A_FILL
        elif priority == 'B':
            p_cell.fill = B_FILL
        else:
            p_cell.fill = C_FILL
        
        ws2.cell(row=row, column=7, value=b.get('相关性', ''))
        ws2.cell(row=row, column=8, value=b.get('url', ''))
    
    # 调整列宽
    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions["B"].width = 40
    
    # 保存
    excel_filename = os.path.join(DATA_DIR, f"bid_report_W{week_num:02d}_{year}_{datetime.now().strftime('%Y%m%d')}.xlsx")
    wb.save(excel_filename)
    print(f"[完成] Excel周报: {excel_filename}")
    
    return excel_filename


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="生成投标情报周报")
    parser.add_argument("--format", "-f", choices=['markdown', 'excel', 'both'], default='excel',
                       help="输出格式: markdown/excel/both")
    args = parser.parse_args()
    
    bids = load_latest_enriched()

    today = datetime.now()
    week_num = today.isocalendar()[1]
    year = today.year

    report, year = generate_markdown_report(bids, week_num)
    filenames = save_report(report, year, week_num, args.format)

    print(f"\n=== 生成完成 ===")
    for f in filenames:
        print(f"  {f}")
