#!/usr/bin/env python3
"""
深化抓取：抓取CSV中所有链接，按模板格式输出Excel
"""
import urllib.request
import urllib.parse
import json
import time
import re
import csv
from datetime import datetime

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Language": "zh-CN,zh;q=0.9",
}

def fetch(url, timeout=20):
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        resp = urllib.request.urlopen(req, timeout=timeout)
        c = resp.read().decode('utf-8', errors='replace')
        resp.close()
        return c
    except:
        return None

def extract(text, patterns):
    for p in patterns:
        m = re.search(p, text, re.DOTALL)
        if m:
            val = m.group(1).strip()
            if val:
                return val[:150]
    return ""

def parse_page(url):
    html = fetch(url)
    if not html or len(html) < 200:
        return None
    
    text = re.sub(r'<[^>]+>', ' ', html)
    text = re.sub(r'\s+', ' ', text)
    
    return {
        "url": url,
        "title": extract(html, [r'<title[^>]*>([^<]+)</title>']),
        "company": extract(text, [
            r'招标[人方][：:]\s*([^\s，,。]{2,40})',
            r'采购[人方][：:]\s*([^\s，,。]{2,40})',
            r'([^\s，,。]{5,30})(?:招标|采购)公告',
        ]),
        "budget": extract(text, [
            r'预算[：:]\s*[￥$]?\s*([\d,，.]+\s*(?:万|元|千万|百万|GWh|MWh|kWh|亿)?)',
            r'采购预算[：:]\s*([\d,，.]+\s*(?:万|元|千万)?)',
            r'(?:项目)?金额[：:]\s*([\d,，.]+\s*(?:万|元|千万|亿)?)',
            r'([\d,，.]+\s*(?:亿|万|元))(?:人民币)?',
        ]),
        "bid_deadline": extract(text, [
            r'(?:截止?|截稿)[^0-9]*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)',
            r'投标截止[^0-9]*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)',
            r'开标[^0-9]*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)',
        ]),
        "publish_date": extract(text, [
            r'发布时间[：:]\s*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)',
            r'发布日期[：:]\s*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)',
            r'发表[于在][：:]\s*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)',
        ]),
        "contact": extract(text, [
            r'联系人[：:]\s*([^\s，,]{2,20})',
            r'电话[：:]\s*([\d\-\s]{7,20})',
            r'联系方式[：:]\s*([^\n\r]{5,50})',
        ]),
    }

def main():
    # 读取CSV中的所有链接
    csv_path = r"C:\Users\16323\Desktop\file_260418_175316_56409.csv"
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        csv_rows = list(reader)
    
    all_links = []
    for row in csv_rows:
        link = row.get("链接", "").strip()
        if link:
            all_links.append({
                "link": link,
                "original": row
            })
    
    print(f"CSV共 {len(all_links)} 条链接\n")
    
    # 抓取每个链接
    results = []
    for i, item in enumerate(all_links, 1):
        link = item["link"]
        orig = item["original"]
        
        short = "/".join(link.split("/")[-2:])
        print(f"[{i}/{len(all_links)}] {short}")
        
        parsed = parse_page(link)
        if parsed:
            results.append({
                "序号": i,
                "优先级": orig.get("优先级", ""),
                "公司名称/主体": parsed.get("company", "") or orig.get("公司名称/主体", ""),
                "项目名称": parsed.get("title", "") or orig.get("项目名称", ""),
                "方向": orig.get("方向", ""),
                "时间": parsed.get("publish_date", "") or parsed.get("bid_deadline", "") or orig.get("时间", ""),
                "预算/金额": parsed.get("budget", "") or orig.get("预算/金额", ""),
                "公司背景": orig.get("公司背景", ""),
                "是否有历史采购/采购动作线索": orig.get("是否有历史采购/采购动作线索", ""),
                "历史采购供应商（已知）": orig.get("历史采购供应商（已知）", ""),
                "机会判断": orig.get("机会判断", ""),
                "链接": link,
            })
            print(f"  OK | {parsed.get('title','')[:40]}")
            print(f"      预算:{parsed.get('budget','')[:20]} 公司:{parsed.get('company','')[:20]}")
        else:
            # 页面抓不到，保留原CSV数据
            results.append({
                "序号": i,
                "优先级": orig.get("优先级", ""),
                "公司名称/主体": orig.get("公司名称/主体", ""),
                "项目名称": orig.get("项目名称", ""),
                "方向": orig.get("方向", ""),
                "时间": orig.get("时间", ""),
                "预算/金额": orig.get("预算/金额", ""),
                "公司背景": orig.get("公司背景", ""),
                "是否有历史采购/采购动作线索": orig.get("是否有历史采购/采购动作线索", ""),
                "历史采购供应商（已知）": orig.get("历史采购供应商（已知）", ""),
                "机会判断": orig.get("机会判断", ""),
                "链接": link,
            })
            print(f"  FAIL (保留原数据)")
        
        time.sleep(0.3)
    
    # 保存JSON
    json_out = f"C:\\Users\\16323\\.openclaw\\workspace\\data\\bid_full_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    with open(json_out, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    # 保存Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "投标机会点"
        
        headers = list(results[0].keys())
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        priority_colors = {
            "S": "FFFF00", "A+": "FF9900", "A": "FFCC00",
            "A-": "CCFF00", "B+": "99FF00", "B": "00FFFF"
        }
        
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
        
        for row_idx, row in enumerate(results, 2):
            priority = row.get("优先级", "")
            
            for col_idx, h in enumerate(headers, 1):
                val = row.get(h, "")
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.alignment = cell_align
                c.border = thin_border
                
                if h == "优先级" and priority in priority_colors:
                    c.fill = PatternFill(
                        start_color=priority_colors[priority],
                        end_color=priority_colors[priority],
                        fill_type="solid"
                    )
                    c.font = Font(bold=True)
        
        col_widths = [6, 8, 25, 40, 20, 14, 18, 25, 28, 25, 35, 65]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        for r in range(2, len(results) + 2):
            ws.row_dimensions[r].height = 40
        
        ws.freeze_panes = "A2"
        
        excel_out = r"C:\Users\16323\Desktop\投标机会点_抓取结果.xlsx"
        wb.save(excel_out)
        print(f"\nExcel: {excel_out}")
    except Exception as e:
        print(f"Excel保存失败: {e}")
    
    print(f"\n完成! 抓取 {len(results)} 条")
    print(f"JSON: {json_out}")

if __name__ == "__main__":
    main()
