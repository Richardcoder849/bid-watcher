#!/usr/bin/env python3
"""抓取CSV模板里的已有链接，生成Excel - 含时间列"""
import urllib.request
import ssl
import re
import json
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "zh-CN,zh;q=0.9",
}

CSV_LINKS = [
    ("https://www.escn.com.cn/news/show-2070171.html", "能源界网", "储能系统设备采购"),
    ("https://eraes.com.cn/newsinfo/9000663.html", "新能源网", "储能框采"),
    ("https://www.cpem.org.cn/list35/117357.html", "电力设备网", "储能EPC"),
    ("https://finance.sina.com.cn/roll/2026-02-25/doc-inhnzrtk5110032.shtml", "新浪财经", "液流储能框采"),
    ("https://www.inengyuan.com/chuneng/12982.html", "能源界", "液冷储能部件框采"),
    ("https://finance.sina.com.cn/roll/2025-02-14/doc-inekmitf4256466.shtml", "新浪财经", "储能系统集采"),
    ("https://m.chu21.com/html/chunengy-22775.shtml", "储能网", "储能EPC"),
    ("https://pvyuan.com/ztb_details?id=469", "pv电站网", "储能EPC"),
]

def fetch_detail(url):
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        resp = urllib.request.urlopen(req, timeout=15, context=SSL_CTX)
        html = resp.read().decode("utf-8", errors="replace")
        resp.close()

        if len(html) < 500:
            return None

        text = re.sub(r"<[^>]+>", " ", html)
        text = re.sub(r"\s+", " ", text).strip()

        def extract(patterns):
            for p in patterns:
                m = re.search(p, text, re.DOTALL)
                if m:
                    val = m.group(1).strip()
                    if val:
                        return val[:100]
            return ""

        title_m = re.search(r"<title>([^<]+)</title>", html)
        title = title_m.group(1).strip() if title_m else ""

        return {
            "url": url,
            "title": title,
            "budget": extract([
                r"预算[：:]\s*[￥$]?\s*([\d,，.]+\s*(?:万|元|千万|亿)?)",
                r"采购预算[：:]\s*([\d,，.]+\s*(?:万|元)?)",
                r"([\d,，.]+\s*(?:亿|万|元))(?:人民币)?",
                r"(\d+\.\d+\s*元/Wh)",
            ]),
            "deadline": extract([
                r"(?:截止?|截稿)[^0-9]*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)",
                r"投标截止[^0-9]*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)",
            ]),
            "company": extract([
                r"招标[人方][：:]\s*([^\s，,。]{2,40})",
                r"采购[人方][：:]\s*([^\s，,。]{2,40})",
            ]),
        }
    except Exception as e:
        return None

def main():
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"[{now}] 抓取 {len(CSV_LINKS)} 个链接...\n")

    results = []
    for i, (url, platform, direction) in enumerate(CSV_LINKS, 1):
        short = url.split("/")[-1][:40]
        print(f"[{i}/{len(CSV_LINKS)}] {short[:40]}")

        detail = fetch_detail(url)
        if detail:
            detail["platform"] = platform
            detail["direction"] = direction
            detail["found_at"] = now
            results.append(detail)

            title = detail.get("title", "无标题")
            budget = detail.get("budget", "无预算")
            print(f"  OK | {title[:40] if title else '无标题'}")
            print(f"      预算: {budget[:20]}")
        else:
            print(f"  FAIL")
            results.append({
                "url": url,
                "platform": platform,
                "direction": direction,
                "title": "",
                "budget": "",
                "deadline": "",
                "company": "",
                "found_at": now,
            })

        time.sleep(1)

    # 生成 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投标机会点"

    headers = ["来源平台", "方向", "标题", "招标单位", "预算/金额", "投标截止", "抓取时间", "链接"]
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    for row_idx, d in enumerate(results, 2):
        row_data = [
            d.get("platform", ""),
            d.get("direction", ""),
            d.get("title", ""),
            d.get("company", ""),
            d.get("budget", ""),
            d.get("deadline", ""),
            d.get("found_at", ""),
            d.get("url", ""),
        ]
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = cell_align
            c.border = thin_border

    col_widths = [12, 15, 40, 25, 20, 18, 16, 70]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    out = r"C:\Users\16323\Desktop\投标机会点_行业平台.xlsx"
    wb.save(out)

    has_budget = sum(1 for d in results if d.get("budget", ""))
    print(f"\n完成! 共 {len(results)} 条")
    print(f"有预算: {has_budget}/{len(results)}")
    print(f"Excel: {out}")

    return results

if __name__ == "__main__":
    main()
