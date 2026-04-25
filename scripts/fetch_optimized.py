#!/usr/bin/env python3
"""
针对性修复失败链接的优化版抓取
"""
import urllib.request
import urllib.parse
import json
import time
import re
import ssl
import csv
from datetime import datetime

# SSL context 忽略证书验证
SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

HEADERS_PC = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Referer": "https://www.bjx.com.cn/",
}

HEADERS_MOBILE = {
    "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1",
    "Accept": "text/html",
    "Accept-Language": "zh-CN,zh;q=0.9",
}

def fetch(url, headers=None, timeout=20, verify_ssl=True, retries=2):
    if headers is None:
        headers = HEADERS_PC
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=headers)
            if verify_ssl:
                resp = urllib.request.urlopen(req, timeout=timeout)
            else:
                resp = urllib.request.urlopen(req, timeout=timeout, context=SSL_CTX)
            c = resp.read().decode('utf-8', errors='replace')
            resp.close()
            return c
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(1)
            else:
                return None
    return None

def is_real_content(html):
    """检查是否返回了真实内容（而非反爬CSS/JS）"""
    if not html or len(html) < 500:
        return False
    text = re.sub(r'<[^>]+>', ' ', html)
    text = re.sub(r'\s+', ' ', text).strip()
    # 如果提取纯文本后少于100个可见字符，可能是反爬
    readable = re.sub(r'[^a-zA-Z0-9\u4e00-\u9fff]', '', text)
    return len(readable) > 100

def extract(text, patterns):
    for p in patterns:
        m = re.search(p, text, re.DOTALL)
        if m:
            val = m.group(1).strip()
            if val:
                return val[:150]
    return ""

def parse_page(url, attempt_mobile=False):
    """解析页面，带多种降级策略"""
    
    # 策略1: 直接抓 PC 版（北极星移动链接 → 尝试去掉 m. 前缀）
    if 'm.bjx.com.cn' in url:
        pc_url = url.replace('https://m.bjx.com.cn', 'https://www.bjx.com.cn')
        html = fetch(pc_url, HEADERS_PC)
        if html and is_real_content(html):
            return parse_html(pc_url, html, "北极星PC")
        
        # 策略2: 试试不www的
        pc_url2 = url.replace('https://m.bjx.com.cn', 'https://bjx.com.cn')
        html = fetch(pc_url2, HEADERS_PC)
        if html and is_real_content(html):
            return parse_html(pc_url2, html, "北极星PC2")
        
        # 策略3: 还是用原URL，但换mobile UA
        html = fetch(url, HEADERS_MOBILE)
        if html and is_real_content(html):
            return parse_html(url, html, "北极星移动UA")
        
        return None
    
    # SSL 证书问题站
    if 'm.chu21.com' in url:
        html = fetch(url, HEADERS_PC, verify_ssl=False)
        if html and is_real_content(html):
            return parse_html(url, html, "chu21(SSL跳过)")
        return None
    
    # 正常站
    html = fetch(url, HEADERS_PC)
    if html and is_real_content(html):
        return parse_html(url, html, "正常")
    
    # 重试带延迟
    time.sleep(2)
    html = fetch(url, HEADERS_PC)
    if html and is_real_content(html):
        return parse_html(url, html, "重试成功")
    
    return None

def parse_html(url, html, source_tag=""):
    text = re.sub(r'<[^>]+>', ' ', html)
    text = re.sub(r'\s+', ' ', text)
    
    return {
        "url": url,
        "title": extract(html, [r'<title[^>]*>([^<]+)</title>']),
        "company": extract(text, [
            r'招标[人方][：:]\s*([^\s，,。]{2,40})',
            r'采购[人方][：:]\s*([^\s，,。]{2,40})',
            r'([^\s，,。]{5,30})(?:招标|采购)公告',
        ]),
        "budget": extract(text, [
            r'预算[：:]\s*[￥$]?\s*([\d,，.]+\s*(?:万|元|千万|百万|GWh|MWh|kWh|亿)?)',
            r'采购预算[：:]\s*([\d,，.]+\s*(?:万|元|千万)?)',
            r'(?:项目)?金额[：:]\s*([\d,，.]+\s*(?:万|元|千万|亿)?)',
            r'([\d,，.]+\s*(?:亿|万|元))(?:人民币)?',
        ]),
        "bid_deadline": extract(text, [
            r'(?:截止?|截稿)[^0-9]*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)',
            r'投标截止[^0-9]*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)',
            r'开标[^0-9]*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)',
        ]),
        "publish_date": extract(text, [
            r'发布时间[：:]\s*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)',
            r'发布日期[：:]\s*(\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}[日]?)',
        ]),
        "contact": extract(text, [
            r'联系人[：:]\s*([^\s，,]{2,20})',
            r'电话[：:]\s*([\d\-\s]{7,20})',
        ]),
        "source": source_tag,
        "found_at": datetime.now().strftime("%Y-%m-%d %H:%M")
    }

def main():
    csv_path = r"C:\Users\16323\Desktop\file_260418_175316_56409.csv"
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        csv_rows = list(csv.DictReader(f))
    
    results = []
    success = 0
    fail = 0
    
    print(f"共 {len(csv_rows)} 条链接，开始优化抓取...\n")
    
    for i, orig in enumerate(csv_rows, 1):
        link = orig.get("链接", "").strip()
        if not link:
            continue
        
        short = link.split('/')[-1][:40]
        print(f"[{i}/{len(csv_rows)}] {short}")
        
        parsed = parse_page(link)
        
        if parsed:
            results.append({
                "序号": i,
                "优先级": orig.get("优先级", ""),
                "公司名称/主体": parsed.get("company", "") or orig.get("公司名称/主体", ""),
                "项目名称": parsed.get("title", "") or orig.get("项目名称", ""),
                "方向": orig.get("方向", ""),
                "时间": parsed.get("publish_date", "") or parsed.get("bid_deadline", "") or orig.get("时间", ""),
                "预算/金额": parsed.get("budget", "") or orig.get("预算/金额", ""),
                "公司背景": orig.get("公司背景", ""),
                "是否有历史采购/采购动作线索": orig.get("是否有历史采购/采购动作线索", ""),
                "历史采购供应商（已知）": orig.get("历史采购供应商（已知）", ""),
                "机会判断": orig.get("机会判断", ""),
                "链接": link,
                "抓取来源": parsed.get("source", ""),
            })
            title = parsed.get('title', '')[:35] if parsed.get('title', '') else '无标题'
            budget = parsed.get('budget', '')[:18] if parsed.get('budget', '') else '无预算'
            src = parsed.get('source', '')
            print(f"  OK [{src}] | {title}")
            print(f"       预算:{budget}")
            success += 1
        else:
            results.append({
                "序号": i,
                "优先级": orig.get("优先级", ""),
                "公司名称/主体": orig.get("公司名称/主体", ""),
                "项目名称": orig.get("项目名称", ""),
                "方向": orig.get("方向", ""),
                "时间": orig.get("时间", ""),
                "预算/金额": orig.get("预算/金额", ""),
                "公司背景": orig.get("公司背景", ""),
                "是否有历史采购/采购动作线索": orig.get("是否有历史采购/采购动作线索", ""),
                "历史采购供应商（已知）": orig.get("历史采购供应商（已知）", ""),
                "机会判断": orig.get("机会判断", ""),
                "链接": link,
                "抓取来源": "原CSV",
            })
            print(f"  FAIL (保留原数据)")
            fail += 1
        
        time.sleep(0.5)
    
    # 保存Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "投标机会点"
        
        headers = list(results[0].keys())
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        priority_colors = {
            "S": "FF0000", "A+": "FF6600", "A": "FFCC00",
            "A-": "CCFF00", "B+": "00FF00", "B": "00FFFF"
        }
        
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
        
        for row_idx, row in enumerate(results, 2):
            priority = row.get("优先级", "")
            for col_idx, h in enumerate(headers, 1):
                val = row.get(h, "")
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.alignment = cell_align
                c.border = thin_border
                if h == "优先级" and priority in priority_colors:
                    c.fill = PatternFill(
                        start_color=priority_colors[priority],
                        end_color=priority_colors[priority],
                        fill_type="solid"
                    )
                    c.font = Font(bold=True)
        
        col_widths = [6, 8, 25, 40, 20, 14, 18, 25, 28, 25, 35, 65, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        for r in range(2, len(results) + 2):
            ws.row_dimensions[r].height = 42
        
        ws.freeze_panes = "A2"
        
        excel_out = r"C:\Users\16323\Desktop\投标机会点_抓取结果_优化版.xlsx"
        wb.save(excel_out)
        print(f"\n{'='*50}")
        print(f"Excel: {excel_out}")
    except Exception as e:
        print(f"Excel失败: {e}")
    
    # 保存JSON
    json_out = f"C:\\Users\\16323\\.openclaw\\workspace\\data\\bid_full_optimized_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    with open(json_out, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\n完成! 成功:{success} 失败:{fail}")
    print(f"JSON: {json_out}")

if __name__ == "__main__":
    main()
